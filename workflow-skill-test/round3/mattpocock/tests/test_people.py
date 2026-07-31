"""人员名单:采购负责人自己维护的一张表,CSV 或 XLSX 都能填。"""

from __future__ import annotations

from pathlib import Path

import pytest

from procurement_qa.model import Role
from procurement_qa.people import RosterError, load_roster

HEADER = ["工号", "姓名", "角色", "负责品类", "联系方式"]
ROWS = [
    ["1001", "李强", "采购员", "", "lq@example.com"],
    ["2001", "王敏", "品类经理", "金属;紧固件", "wm@example.com"],
    ["2002", "赵磊", "品类经理", "包材", "zl@example.com"],
    ["9001", "陈总", "品类经理", "", "cz@example.com"],
]


def _write_csv(path: Path, rows: list[list[str]]) -> None:
    import csv

    with path.open("w", encoding="utf-8", newline="") as fh:
        csv.writer(fh).writerows(rows)


def _write_xlsx(path: Path, rows: list[list[str]]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    for row in rows:
        wb.active.append(row)
    wb.save(str(path))


@pytest.fixture(params=["csv", "xlsx"])
def roster_path(request, tmp_path: Path) -> Path:
    path = tmp_path / f"名单.{request.param}"
    rows = [HEADER, *ROWS]
    (_write_csv if request.param == "csv" else _write_xlsx)(path, rows)
    return path


def test_CSV和XLSX读出来的名单一模一样(roster_path: Path) -> None:
    roster = load_roster(roster_path, fallback_employee_id="9001")

    assert [a.employee_id for a in roster.askers] == ["1001", "2001", "2002", "9001"]
    assert roster.find("1001").name == "李强"
    assert roster.find("1001").role is Role.BUYER
    assert roster.find("2001").role is Role.CATEGORY_MANAGER


def test_一个人可以负责多个品类(roster_path: Path) -> None:
    roster = load_roster(roster_path, fallback_employee_id="9001")

    assert roster.find("2001").categories == ("金属", "紧固件")
    assert roster.find("2002").categories == ("包材",)


def test_查不到工号是明确的错误而不是当成采购员放行(roster_path: Path) -> None:
    roster = load_roster(roster_path, fallback_employee_id="9001")

    with pytest.raises(RosterError, match="7777"):
        roster.find("7777")


def test_按品类找得到对应的品类经理(roster_path: Path) -> None:
    roster = load_roster(roster_path, fallback_employee_id="9001")

    assert roster.manager_of("金属").name == "王敏"
    assert roster.manager_of("包材").name == "赵磊"


def test_没人认领的品类落到兜底接收人(roster_path: Path) -> None:
    roster = load_roster(roster_path, fallback_employee_id="9001")

    assert roster.manager_of("化学品") is None
    assert roster.fallback.name == "陈总"


def test_角色只认采购员和品类经理别的值直接报错(tmp_path: Path) -> None:
    path = tmp_path / "名单.csv"
    _write_csv(path, [HEADER, ["1001", "李强", "实习生", "", ""]])

    with pytest.raises(RosterError, match="实习生"):
        load_roster(path, fallback_employee_id="1001")


def test_兜底接收人不在名单里也要报错(roster_path: Path) -> None:
    with pytest.raises(RosterError, match="兜底"):
        load_roster(roster_path, fallback_employee_id="8888")


def test_XLSX里有多个有内容的工作表要报错而不是闷头读第一个(tmp_path: Path) -> None:
    """出自 code-review 的 Spec #3:静默读错表和本项目"绝不静默丢弃"的基调冲突。"""
    from openpyxl import Workbook

    path = tmp_path / "名单.xlsx"
    wb = Workbook()
    for row in [HEADER, *ROWS]:
        wb.active.append(row)
    second = wb.create_sheet("上季度名单")
    for row in [HEADER, ["1001", "旧的李强", "采购员", "", ""]]:
        second.append(row)
    wb.save(str(path))

    with pytest.raises(RosterError, match="不止一个"):
        load_roster(path, fallback_employee_id="9001")
