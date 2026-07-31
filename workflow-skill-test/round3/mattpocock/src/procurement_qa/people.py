"""人员名单:系统判断提问人是谁、能看什么、转给谁的唯一依据。

CSV 与 XLSX 走同一个接口 —— 采购负责人用 Excel 填,存成哪种都行。
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from .model import Asker, Role

COLUMNS = ("工号", "姓名", "角色", "负责品类", "联系方式")
_CATEGORY_SEPARATORS = ";;,、,"


class RosterError(Exception):
    """名单本身有问题,或者名单里查不到这个人。"""


@dataclass(frozen=True)
class Roster:
    askers: tuple[Asker, ...]
    fallback: Asker

    def find(self, employee_id: str) -> Asker:
        for asker in self.askers:
            if asker.employee_id == str(employee_id).strip():
                return asker
        raise RosterError(f"人员名单里没有工号 {employee_id},无法确认身份")

    def manager_of(self, category: str) -> Asker | None:
        for asker in self.askers:
            if asker.manages(category):
                return asker
        return None


def load_roster(path: Path | str, *, fallback_employee_id: str) -> Roster:
    path = Path(path)
    rows = _read_xlsx(path) if path.suffix.lower() in {".xlsx", ".xlsm"} else _read_csv(path)
    if not rows:
        raise RosterError(f"人员名单 {path} 是空的")

    header = [str(c or "").strip() for c in rows[0]]
    missing = [c for c in COLUMNS[:4] if c not in header]
    if missing:
        raise RosterError(f"人员名单缺少必需的列:{'、'.join(missing)}(需要 {'、'.join(COLUMNS)})")
    index = {name: header.index(name) for name in COLUMNS if name in header}

    askers: list[Asker] = []
    for lineno, row in enumerate(rows[1:], start=2):
        cells = [str(c or "").strip() for c in row]
        if not any(cells):
            continue

        def cell(name: str) -> str:
            pos = index.get(name, -1)
            return cells[pos] if 0 <= pos < len(cells) else ""

        employee_id = cell("工号")
        if not employee_id:
            raise RosterError(f"人员名单第 {lineno} 行缺工号")
        askers.append(
            Asker(
                employee_id=employee_id,
                name=cell("姓名"),
                role=_parse_role(cell("角色"), lineno),
                categories=_parse_categories(cell("负责品类")),
                contact=cell("联系方式"),
            )
        )

    ids = [a.employee_id for a in askers]
    if len(set(ids)) != len(ids):
        raise RosterError("人员名单里有重复的工号")

    fallback_id = str(fallback_employee_id).strip()
    fallback = next((a for a in askers if a.employee_id == fallback_id), None)
    if fallback is None:
        raise RosterError(f"兜底接收人 {fallback_employee_id} 不在人员名单里")

    return Roster(askers=tuple(askers), fallback=fallback)


def _parse_role(raw: str, lineno: int) -> Role:
    for role in Role:
        if raw == role.value:
            return role
    raise RosterError(
        f"人员名单第 {lineno} 行的角色「{raw}」不认识,只能是「{Role.BUYER.value}」或「{Role.CATEGORY_MANAGER.value}」"
    )


def _parse_categories(raw: str) -> tuple[str, ...]:
    if not raw:
        return ()
    normalised = raw
    for sep in _CATEGORY_SEPARATORS[1:]:
        normalised = normalised.replace(sep, _CATEGORY_SEPARATORS[0])
    return tuple(part.strip() for part in normalised.split(_CATEGORY_SEPARATORS[0]) if part.strip())


def _read_csv(path: Path) -> list[list[str]]:
    import csv

    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return [list(row) for row in csv.reader(fh)]


def _read_xlsx(path: Path) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=True, read_only=True)
    populated = [ws for ws in wb.worksheets if ws.max_row and ws.max_row > 1]
    if len(populated) > 1:
        wb.close()
        raise RosterError(
            f"人员名单 {path.name} 里有不止一个有内容的工作表"
            f"({'、'.join(ws.title for ws in populated)}),请只保留一个,免得读错表。"
        )
    rows = [
        ["" if v is None else str(v) for v in row]
        for row in wb.active.iter_rows(values_only=True)
    ]
    wb.close()
    return rows
